import sqlite3
import math
import time
import openpyxl
from datetime import datetime
from datetime import date
from flask import flash
 
class FDataBase:
    def __init__(self, db):
        self.__db = db
        self.__cur = db.cursor()
 
    def getMenu(self):
        sql = '''SELECT * FROM mainmenu'''
        try:
            self.__cur.execute(sql)
            res = self.__cur.fetchall()
            if res: return res
        except:
            print("Ошибка чтения из БД")
        return []

    def addrequest(self, pp, last_name, 
    first_name, o_patronymic, d_position, data_nach, data_kon, birthday):
        try:
            data_t = date.today()
            self.__cur.execute("INSERT INTO employee VALUES(NULL, ?, ?, ?, ?, ?)", (last_name, 
    first_name, o_patronymic, d_position, birthday))
            self.__cur.execute("INSERT INTO addrequest VALUES(NULL, ?, ?, ?, ?, ?)", (pp, data_nach, 
            data_kon, data_t, 1))
            self.__db.commit()
        except sqlite3.Error as e:
            print("Ошибка добавления запроса в БД "+str(e))
            return False
 
        return True

    def getInfoEmployeesRequests(self, requestId):
        try:
            self.__cur.execute(f"SELECT last_name, first_name, o_patronymic, d_position FROM employee WHERE id = {requestId} LIMIT 1")
            res = self.__cur.fetchone()
            if res:
                return res
        except sqlite3.Error as e:
            print("Ошибка получения запроса из БД "+str(e))
        return (False, False)

    def getAllEmployee(self):
        try:
            self.__cur.execute(f"SELECT id, last_name, first_name, o_patronymic, pp FROM allemployee")
            res = self.__cur.fetchall()
            if res:
                return res
        except sqlite3.Error as e:
            print("Ошибка получения запроса из БД "+str(e))
        return []

    def getInfoAddRequests(self, eployees_id):
        try:
            self.__cur.execute(f"SELECT pp, data_nach, data_kon, data_tek FROM addrequest WHERE employee_id  = {eployees_id} LIMIT 1")
            res = self.__cur.fetchone()
            if res:
                return res
        except sqlite3.Error as e:
            print("Ошибка получения запроса из БД "+str(e))
        return (False, False)

    def getRequestsAnons(self):
        try:
            self.__cur.execute(f"SELECT id, last_name, first_name, o_patronymic, d_position FROM employee")
            res = self.__cur.fetchall()
            if res: return res
        except sqlite3.Error as e:
            print("Ошибка получения запроса из БД "+str(e))
 
        return []
        
    def addUser(self, login, hpsw):
        try:
            self.__cur.execute(f"SELECT COUNT() as `count` FROM users WHERE login LIKE '{login}'")
            res = self.__cur.fetchone()
            if res['count'] > 0:
                print("Пользователь с таким логином уже существует")
                return False
            tm = math.floor(time.time())
            self.__cur.execute("INSERT INTO users VALUES (NULL, ?, ?, ?)", (login, hpsw, tm))
            self.__db.commit()
        except sqlite3.Error as e:
            print("Ошибка добавления в БД" + str(e))
            return False
        return True

    def getUser(self, user_id):
        try:
            self.__cur.execute(f"SELECT * FROM users WHERE id = {user_id} LIMIT 1")
            res = self.__cur.fetchone()
            if not res:
                print("Пользователь не найден")
                return False
            return res
        except sqlite3.Error as e:
            print("Ошибка при обращении к БД" + str(e))
        return False


    def getUserByLogin(self, login):
        try:
            self.__cur.execute(f"SELECT * FROM users WHERE login = '{login}' LIMIT 1")
            res = self.__cur.fetchone()
            if not res:
                print("Пользователь не найден")
                return False
            return res

        except sqlite3.Error as e:
            print("Ошибка добавления в БД" + str(e))

        return False

    def getDocumentsAnons(self):
        try:
            self.__cur.execute(f"SELECT id, inventory_number, name_doc, type_doc, pp, years, shelf_life, counts, data_add FROM documents")
            res = self.__cur.fetchall()
            if res: return res
        except sqlite3.Error as e:
            print("Ошибка получения документов из БД "+str(e))
 
        return []   


    def getDocumentsDeleteAnons(self):
        try:
            self.__cur.execute(f"SELECT id, del_number, name_doc, type_doc, pp, years,  shelf_life, counts, data_add, data_del FROM deletes ")
            res = self.__cur.fetchall()
            if res: return res
        except sqlite3.Error as e:
            print("Ошибка получения документов из БД "+str(e))
 
        return []   
    

    def getOrdersAnons(self, eployees_id):
        try:
            self.__cur.execute(f"SELECT id, employee_id, documents_id, type_orders, data, nomber, page FROM orders WHERE employee_id  = {eployees_id}")
            res = self.__cur.fetchall()
            if res: return res
            else:
                return ('Приказов на данного сотрудника нет в базе')
        except sqlite3.Error as e:
            print("Ошибка получения документов из БД "+str(e))
 
        return [] 
    def getListOrders(self):
        try:
            self.__cur.execute(f"SELECT id, employee_id, documents_id, type_orders, data, nomber, page FROM orders")
            res = self.__cur.fetchall()
            if res: return res
            else:
                return ('Приказов на данного сотрудника нет в базе')
        except sqlite3.Error as e:
            print("Ошибка получения документов из БД "+str(e))

    def getProfileAnons(self, eployees_id):
        try:
            self.__cur.execute(f"SELECT id, years FROM personal_accounts WHERE employee_id  = {eployees_id}")
            res = self.__cur.fetchall()
            if res: return res
            else:
                return ('Лицевых счетов на данного сотрудника нет в базе')
        except sqlite3.Error as e:
            print("Ошибка получения документов из БД "+str(e))
        return (False,False)

    def getCalc_MonthAnons(self, allprofile):
        try:
            allres = []
            for profile in allprofile:
                self.__cur.execute(f"SELECT id, personal_accounts_id, month_calc, summ FROM calc_by_month WHERE personal_accounts_id  = {profile[0]}")
                res = self.__cur.fetchall()
                allres += res
            if allres: return allres
            else:
                return ('Лицевых счетов на данного сотрудника нет в базе')
        except sqlite3.Error as e:
            print("Ошибка получения документов из БД "+str(e))
        return []
    def getListProfile(self):
        try:
            self.__cur.execute(f"SELECT id, employee_id, years FROM personal_accounts")
            res = self.__cur.fetchall()
            if res: return res
            else:
                return ('Лицевых счетов на данного сотрудника нет в базе')
        except sqlite3.Error as e:
            print("Ошибка получения документов из БД "+str(e))
        return (False,False)

    def getListCalc_Month (self):
        try:
            self.__cur.execute(f"SELECT id, personal_accounts_id, month_calc, summ FROM calc_by_month")
            res = self.__cur.fetchall()
            if res: return res
            else:
                return ('Лицевых счетов на данного сотрудника нет в базе')
        except sqlite3.Error as e:
            print("Ошибка получения документов из БД "+str(e))
        return (False,False) 

    def getListInventories(self, name_faile):
            try:
                book = openpyxl.open(name_faile, read_only= True)
                sheet = book.active
                for r in range(2, sheet.max_row+1):
                    inventory_number = sheet[r][0].value
                    name_doc = sheet[r][1].value
                    type_doc = sheet[r][2].value
                    pp = sheet[r][3].value
                    years = sheet[r][4].value
                    shelf_life = sheet[r][5].value
                    counts = sheet[r][6].value
                    tm = date.today()
                    values = (inventory_number, name_doc, type_doc, pp, years, shelf_life, counts, tm)
                    self.__cur.execute("INSERT INTO documents VALUES (NULL,?,?,?,?,?,?,?,?)", values)
                    self.__db.commit()
            except sqlite3.Error as e:
                flash("Ошибка загрузки файла Excel "+str(e))

    def getExeleOrders (self, name_faile):
        try:
            book = openpyxl.open(name_faile, read_only= True)
            sheet = book.active
            self.__cur.execute(f"SELECT id, last_name,first_name, o_patronymic FROM allemployee") 
            allemployee = self.__cur.fetchall()
            
            for r in range(2, sheet.max_row+1):
                id_e = 0
                F = sheet[r][0].value
                I = sheet[r][1].value
                O = sheet[r][2].value
                pp = sheet[r][3].value
                fio = (F,I,O,pp)
                
                for employee in allemployee:
                    if F == employee[1] and I == employee[2] and O == employee[3]:
                       id_e = employee[0]
                if id_e == 0:
                    self.__cur.execute("INSERT INTO allemployee VALUES (NULL,?,?,?,?)", fio)
                    self.__db.commit()
                    id_e = self.__cur.lastrowid
                number_doc = sheet[r][4].value
                type_orders = sheet[r][5].value
                data = sheet[r][6].value
                nomber = sheet[r][7].value
                page = sheet[r][8].value
                values = (id_e,number_doc, type_orders,data,nomber,page)
                self.__cur.execute("INSERT INTO orders VALUES (NULL,?,?,?,?,?,?)", values)
                self.__db.commit()
        except sqlite3.Error as e:
            flash("Ошибка загрузки файла Excel "+str(e))

    def getExelepersonal_accounts(self,name_faile):
        try:
            book = openpyxl.open(name_faile, read_only= True)
            sheet = book.active
            self.__cur.execute(f"SELECT id, last_name,first_name, o_patronymic FROM allemployee") 
            allemployee = self.__cur.fetchall()
            self.__cur.execute(f"SELECT id FROM personal_accounts") 
            personal_account = self.__cur.fetchall()
            for r in range(2, sheet.max_row+1):
                id_e = 0
                id_p = 0
                F = sheet[r][0].value
                I = sheet[r][1].value
                O = sheet[r][2].value
                pp = sheet[r][4].value
                fio = (F,I,O,pp)
                for employee in allemployee:
                    if F == employee[1] and I == employee[2] and O == employee[3]:
                        id_e = employee[0]
                if id_e == 0:
                    self.__cur.execute("INSERT INTO allemployee VALUES (NULL,?,?,?,?)", fio)
                    self.__db.commit()
                    id_e = self.__cur.lastrowid
                years = sheet[r][3].value
                values = (id_e, years)
                for person in personal_account:
                    if id_p == person[0]:
                        id_p = person[0]
                if id_p == 0:
                    self.__cur.execute("INSERT INTO personal_accounts VALUES (NULL,?,?)", values)
                    self.__db.commit()
                    id_p = self.__cur.lastrowid
                for s in range(6,18):
                    month = sheet[1][s].value
                    summ = sheet[r][s].value
                    values = (id_p, month, summ )
                    self.__cur.execute("INSERT INTO calc_by_month VALUES (NULL,?,?,?)", values)
                    self.__db.commit()

        except sqlite3.Error as e:
            print("Ошибка загрузки файла Excel "+str(e))

    def getExelDocumentsdelite(self, doc_del):
        try:
            row = []
            id = 1
            book = openpyxl.open('ACT.xlsx', read_only= False)
            sheet = book['1'] 
            for doc in doc_del:
                row = [id,doc[2],doc[5],doc[1],doc[7],doc[7],doc[6]]
                sheet.append(row)
                id += 1
            book.save('Акт утилизации.xlsx')
            return ('Акт утилизации.xlsx')
        except sqlite3.Error as e:
            print("Ошибка скачивания файла Excel "+ str(e))

    def DeletAcceptance(self, doc_del):
        try:
            self.__cur.execute(f"SELECT id, inventory_number, name_doc, type_doc, pp, years, shelf_life, counts, data_add FROM documents")
            res = self.__cur.fetchall()
            datas = date.today()
            for doc in doc_del:
                for r in res:
                    if doc[0] == r [0]:
                        self.__cur.execute(f"DELETE FROM documents WHERE id = {doc[0]}")
                        self.__db.commit()
                values = (doc[1],doc[2],doc[3],doc[4],doc[5],doc[6],doc[7],doc[8],datas)
                print(values)
                self.__cur.execute("INSERT INTO deletes VALUES (NULL,?,?,?,?,?,?,?,?,?)", values)
                self.__db.commit()
        except sqlite3.Error as e:
            print("Ошибка удаления описи "+ str(e))


    def getInventories_id(self):
        try:
            self.__cur.execute(f"SELECT id  FROM inventories") 
            res = self.__cur.fetchall()
            if res: return res
        except sqlite3.Error as e:
            print("Ошибка получения документов из БД "+str(e))
        return (False,False)

    def getAllEmployeeSerch(self, serch):
        try:
            self.__cur.execute(f"SELECT id, last_name, first_name, o_patronymic, pp FROM allemployee WHERE last_name = {str(serch)}")
            res = self.__cur.fetchall()
            for r in res:
                print(type(r[1]))
            if res: return res
        except sqlite3.Error as e:
            print("Ошибка получения запроса из БД "+str(e))
        return []
    
    def getDocumentsdelite(self):
        try:
            self.__cur.execute(f"SELECT id, inventory_number, name_doc, type_doc, pp, years, shelf_life, counts, data_add FROM documents")
            res = self.__cur.fetchall()
            delet = []
            dates = date.today()
            for r in res:
                summ = r[5] + r[6]
                if summ < dates.year:
                    delet += [r]
            if delet: return delet
        except sqlite3.Error as e:
            print("Ошибка получения документов из БД "+str(e))
 
        return []  
       
                
     